import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

# הגדרות דף
st.set_page_config(page_title="ניהול חלוקת מבחנים", layout="centered")

def apply_design(ws, g_df, start_r, r_id):
    side = Side(style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    ws.cell(row=start_r, column=1, value=f"חדר: {r_id}").font = Font(bold=True, size=14)
    header_row = start_r + 1
    
    for c_idx, col_name in enumerate(g_df.columns, 1):
        cell = ws.cell(row=header_row, column=c_idx, value=col_name)
        cell.font, cell.fill, cell.border = Font(bold=True), header_fill, border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    for r_idx, row_vals in enumerate(g_df.values, 1):
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=header_row + r_idx, column=c_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
    
    return header_row + len(g_df) + 5

st.title("📋 כלי חלוקת תלמידים לחדרים")
st.write("העלה קובץ אקסל, הגדר חדרים והורד את הקובץ המעוצב.")

uploaded_file = st.file_uploader("בחר קובץ אקסל (מתמטיקה.xlsx)", type=["xlsx"])

if uploaded_file:
    # קריאת הקובץ
    df = pd.read_excel(uploaded_file, header=2, engine='calamine').dropna(how='all')
    
    # עיבוד עמודות
    t_cols = ['מס\'', 'שם התלמיד']
    time_cols = [c for c in df.columns if 'תוספת זמן' in str(c)]
    extra_cols = ['הבחנות בחדר מצומצם', 'הגדלת שאלון', 'הקראת שאלון']
    cols = [c for c in t_cols if c in df.columns]
    rem = [c for c in (time_cols + extra_cols) if c in df.columns]
    df_f = df[cols + rem].copy()
    
    if 'שם התלמיד' in df_f.columns:
        pos = df_f.columns.get_loc('שם התלמיד') + 1
        df_f.insert(pos, 'נוכחות', "")
        df_f.insert(pos + 1, 'הגשה', "")

    st.success(f"נמצאו {len(df_f)} תלמידים בקובץ.")
    
    # ממשק חלוקה לחדרים
    st.subheader("חלוקה לחדרים")
    if 'rooms' not in st.session_state:
        st.session_state.rooms = []

    col1, col2 = st.columns(2)
    with col1:
        room_name = st.text_input("מספר חדר/שם:")
    with col2:
        num_students = st.number_input("מספר תלמידים:", min_value=1, max_value=len(df_f), value=10)

    if st.button("הוסף חדר"):
        current_allocated = sum(len(g[1]) for g in st.session_state.rooms)
        if current_allocated + num_students <= len(df_f):
            chunk = df_f.iloc[current_allocated : current_allocated + num_students]
            st.session_state.rooms.append((room_name, chunk))
        else:
            st.error("אין מספיק תלמידים נותרים לחלוקה זו.")

    # הצגת המצב הנוכחי
    allocated_so_far = sum(len(g[1]) for g in st.session_state.rooms)
    st.write(f"תלמידים שחולקו: {allocated_so_far} מתוך {len(df_f)}")
    
    for r_id, r_df in st.session_state.rooms:
        st.info(f"חדר {r_id}: {len(r_df)} תלמידים")

    if st.button("איפוס חלוקה"):
        st.session_state.rooms = []
        st.rerun()

    # יצירת הקובץ להורדה
    if st.session_state.rooms:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame().to_excel(writer, sheet_name='רשימה', index=False)
            ws = writer.sheets['רשימה']
            ws.sheet_view.rightToLeft = True
            
            curr_r = 1
            for r_id, g_df in st.session_state.rooms:
                curr_r = apply_design(ws, g_df, curr_r, r_id)
                
            for column in ws.columns:
                h_val = column[1].value
                ws.column_dimensions[column[0].column_letter].width = 16 if h_val == "שם התלמיד" else 9
        
        st.download_button(
            label="📥 הורד קובץ אקסל מוכן",
            data=output.getvalue(),
            file_name="split_exams.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )